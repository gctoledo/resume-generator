from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

def apply_excel_styles(ws: Worksheet, column_headers: list[str]) -> None:
    fill_header = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border_thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    bold = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill_header
        cell.border = border_thin
        cell.alignment = align_center

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for idx, cell in enumerate(row, start=1):
            cell.border = border_thin
            header = ws.cell(row=1, column=idx).value
            if idx == 2:
                cell.alignment = align_wrap
            else:
                cell.alignment = align_center

            if idx >= 3:
                if cell.value is not None:
                    cell.number_format = 'R$ #,##0.00'
                    cell.fill = fill_green
                else:
                    cell.fill = fill_red

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        header = col[0].value
        if header == "Cliente":
            ws.column_dimensions[col_letter].width = 50
        elif header in column_headers[2:]:
            ws.column_dimensions[col_letter].width = 18
        else:
            max_length = max((len(str(cell.value)) for cell in col if cell.value is not None), default=0)
            ws.column_dimensions[col_letter].width = max_length + 6

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 22

    ws.freeze_panes = "A2"
